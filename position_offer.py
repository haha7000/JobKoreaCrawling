"""
포지션 제안 문구 생성
자기소개서 + 자격증 정보를 기반으로 LLM이 맞춤형 제안 문구 생성
"""
import json
import os
from pathlib import Path
from typing import Optional, Dict, List
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, Alignment


class PositionOfferGenerator:
    """포지션 제안 문구 생성기"""

    # 기본 템플릿
    BASE_TEMPLATE = """안녕하세요. 한국중소기업진흥원 입니다.
저희가 찾고있는 포지션에 적합한 인재라고 생각되어 이렇게 제안 드립니다.
긍정적인 검토 부탁 드리며, 관련 자세한 내용이 궁금하시다면 응답기간 내 회신 부탁 드립니다."""

    def __init__(self, api_key: Optional[str] = None):
        """
        Args:
            api_key: OpenAI API 키 (None이면 환경변수에서 자동 로드)
        """
        self.api_key = api_key or os.getenv("OPENAI_API_KEY_COMPANY")
        if not self.api_key:
            raise ValueError("OpenAI API 키가 필요합니다. 환경변수 OPENAI_API_KEY를 설정하거나 인자로 전달하세요.")

        self.client = OpenAI(api_key=self.api_key)

    def _create_prompt(self, person_data: Dict) -> str:
        """
        LLM에게 전달할 프롬프트 생성

        Args:
            person_data: 개인 정보 (자기소개서, 자격증 포함)

        Returns:
            프롬프트 문자열
        """
        # 자기소개서 요약
        intro_summary = ""
        if person_data.get("자기소개서"):
            intro_list = person_data["자기소개서"]
            if intro_list and len(intro_list) > 0:
                # 첫 번째 자기소개서만 사용 (길이 제한)
                first_intro = intro_list[0].get("body_text", "")
                if first_intro:
                    # 너무 길면 앞부분만 (최대 500자)
                    intro_summary = first_intro[:500]

        # 자격증 요약
        cert_summary = ""
        if person_data.get("자격증"):
            cert_list = person_data["자격증"]
            if cert_list:
                cert_names = [c.get("자격증명", "") for c in cert_list if c.get("자격증명")]
                cert_summary = ", ".join(cert_names[:5])  # 최대 5개만

        # 기본 정보
        career = person_data.get("경력", "")
        job = person_data.get("직무", "")

        prompt = f"""당신은 채용 담당자입니다.
아래 정보를 바탕으로 **기존 포지션 제안 문구를 조금만 수정**하여 맞춤형 제안 문구를 작성해주세요.

## 기존 문구 (이것을 기반으로 수정):
{self.BASE_TEMPLATE}

## 지원자 정보:
- 경력: {career}
- 직무: {job}
- 자기소개서 요약: {intro_summary if intro_summary else "(없음)"}
- 자격증: {cert_summary if cert_summary else "(없음)"}

## 요구사항:
1. **기존 문구의 구조와 톤을 유지**하세요
2. 지원자의 경력, 강점, 자격증을 **자연스럽게 1-2문장만 추가**하세요
3. 너무 길지 않게 (전체 3-4문장)
4. 존댓말 유지
5. "한국중소기업진흥원"이라는 회사명은 유지
6. **지원자의 이름은 절대 언급하지 마세요** (OO님, 귀하 등도 사용 금지)

## 출력 형식:
수정된 제안 문구만 출력하세요. 부연 설명 없이."""

        return prompt

    def generate_offer(self, person_data: Dict) -> str:
        """
        개인 정보를 기반으로 포지션 제안 문구 생성

        Args:
            person_data: 개인 정보 딕셔너리

        Returns:
            생성된 제안 문구
        """
        # 자기소개서나 자격증이 없으면 기본 템플릿 사용
        has_intro = person_data.get("자기소개서") and len(person_data["자기소개서"]) > 0
        has_cert = person_data.get("자격증") and len(person_data["자격증"]) > 0

        if not has_intro and not has_cert:
            print(f"   ⚠️  자기소개서/자격증 없음 - 기본 템플릿 사용")
            return self.BASE_TEMPLATE

        try:
            # 프롬프트 생성
            prompt = self._create_prompt(person_data)

            # OpenAI API 호출
            response = self.client.chat.completions.create(
                model="gpt-4o", 
                messages=[
                    {"role": "system", "content": "당신은 전문 채용 담당자입니다."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.7
            )

            generated_text = response.choices[0].message.content.strip()
            return generated_text

        except Exception as e:
            print(f"   ❌ LLM 생성 오류: {e}")
            return self.BASE_TEMPLATE

    def process_file(self, input_json: str, output_json: str):
        """
        JSON 파일을 읽어서 모든 지원자에 대해 제안 문구 생성

        Args:
            input_json: 입력 JSON 파일 경로 (with_details.json)
            output_json: 출력 JSON 파일 경로
        """
        # 입력 파일 로드
        if not Path(input_json).exists():
            print(f"❌ 파일을 찾을 수 없습니다: {input_json}")
            return

        with open(input_json, 'r', encoding='utf-8') as f:
            people = json.load(f)

        print(f"📋 총 {len(people)}명 처리 시작\n")

        # 각 사람에 대해 제안 문구 생성
        for idx, person in enumerate(people, 1):
            name = person.get("이름", "Unknown")
            print(f"[{idx}/{len(people)}] {name}")

            # 제안 문구 생성
            offer_text = self.generate_offer(person)

            # 결과에 추가
            person["포지션제안문구"] = offer_text

            # 미리보기
            preview = offer_text[:80] + "..." if len(offer_text) > 80 else offer_text
            print(f"   💬 {preview}\n")

            # 즉시 저장 (중간 손실 방지)
            with open(output_json, 'w', encoding='utf-8') as f:
                json.dump(people, f, ensure_ascii=False, indent=2)

        print(f"\n✅ 완료! {len(people)}명의 제안 문구 생성")
        print(f"💾 저장: {output_json}")


def update_excel_with_offers(excel_path: str, candidates: List[Dict], min_score: int = 30):
    """
    엑셀 파일에 제안문구 컬럼 추가 (30점 이상만)

    Args:
        excel_path: 엑셀 파일 경로
        candidates: 제안문구가 포함된 후보자 리스트
        min_score: 최소 점수 (기본 30점)
    """
    # 이력서번호 → 제안문구 매핑
    offer_map = {}
    for candidate in candidates:
        rno = candidate.get("이력서번호")
        score = candidate.get("점수상세", {}).get("총점", 0)
        offer = candidate.get("포지션제안문구", "")

        # 30점 이상만 제안문구 추가
        if rno and score >= min_score and offer:
            offer_map[str(rno)] = offer

    # 엑셀 파일 열기
    if not Path(excel_path).exists():
        print(f"⚠️  엑셀 파일을 찾을 수 없습니다: {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 헤더 찾기
    headers = [cell.value for cell in ws[1]]

    # "제안문구" 컬럼이 이미 있는지 확인
    if "제안문구" in headers:
        print(f"⚠️  '제안문구' 컬럼이 이미 존재합니다. 기존 제안문구를 업데이트합니다.")
        offer_col_idx = headers.index("제안문구") + 1
    else:
        # 맨 뒤에 "제안문구" 컬럼 추가
        offer_col_idx = len(headers) + 1
        ws.cell(row=1, column=offer_col_idx).value = "제안문구"
        ws.cell(row=1, column=offer_col_idx).font = Font(bold=True)
        ws.cell(row=1, column=offer_col_idx).alignment = Alignment(horizontal='center')

        # 헤더 다시 읽기
        headers = [cell.value for cell in ws[1]]

    # 이력서번호 컬럼 위치 찾기
    rno_col_idx = None
    if "이력서번호" in headers:
        rno_col_idx = headers.index("이력서번호") + 1

    if not rno_col_idx:
        print(f"⚠️  '이력서번호' 컬럼을 찾을 수 없습니다.")
        return

    # 각 행에 제안문구 입력 (30점 이상만)
    updated_count = 0
    for row_idx in range(2, ws.max_row + 1):
        rno = ws.cell(row=row_idx, column=rno_col_idx).value
        if rno and str(rno) in offer_map:
            offer = offer_map[str(rno)]
            ws.cell(row=row_idx, column=offer_col_idx).value = offer
            ws.cell(row=row_idx, column=offer_col_idx).alignment = Alignment(wrap_text=True, vertical='top')
            updated_count += 1

    # 제안문구 컬럼 너비 조정
    ws.column_dimensions[openpyxl.utils.get_column_letter(offer_col_idx)].width = 50

    # 저장
    wb.save(excel_path)
    print(f"\n📊 엑셀 파일 업데이트 완료!")
    print(f"   파일: {excel_path}")
    print(f"   제안문구 입력: {updated_count}개 행 ({min_score}점 이상만)")


def main():
    """메인 실행"""
    # 설정
    INPUT_FILE = "output/kspac2022_with_introduction.json"  # Detail.py의 출력 (전체)
    OUTPUT_FILE = "output/kspac2022_with_offers.json"
    EXCEL_FILE = "output/kspac2022_결과.xlsx"
    MIN_SCORE = 30

    # OpenAI API 키 설정 (환경변수 또는 직접 입력)
    # export OPENAI_API_KEY_COMPANY="sk-..."
    # 또는
    # API_KEY = "sk-..."

    try:
        # 1. 입력 파일 로드
        if not Path(INPUT_FILE).exists():
            print(f"❌ 파일을 찾을 수 없습니다: {INPUT_FILE}")
            return

        with open(INPUT_FILE, 'r', encoding='utf-8') as f:
            candidates = json.load(f)

        print(f"📋 총 {len(candidates)}명 로드")

        # 2. 30점 이상인 사람만 필터링
        qualified = []
        for candidate in candidates:
            score = candidate.get("점수상세", {}).get("총점", 0)
            if score >= MIN_SCORE:
                qualified.append(candidate)

        print(f"🎯 {MIN_SCORE}점 이상: {len(qualified)}명")
        print(f"   → 제안문구 생성 대상\n")

        if len(qualified) == 0:
            print(f"⚠️  {MIN_SCORE}점 이상인 사람이 없습니다.")
            return

        # 3. 제안문구 생성
        generator = PositionOfferGenerator()

        for idx, candidate in enumerate(qualified, 1):
            name = candidate.get("이름", "Unknown")
            score = candidate.get("점수상세", {}).get("총점", 0)
            print(f"[{idx}/{len(qualified)}] {name} ({score}점)")

            # 제안 문구 생성
            offer_text = generator.generate_offer(candidate)
            candidate["포지션제안문구"] = offer_text

            # 미리보기
            preview = offer_text[:80] + "..." if len(offer_text) > 80 else offer_text
            print(f"   💬 {preview}\n")

        # 4. JSON 저장 (합격자만)
        with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
            json.dump(qualified, f, ensure_ascii=False, indent=2)

        print(f"✅ 제안문구 생성 완료!")
        print(f"💾 저장: {OUTPUT_FILE}\n")

        # 5. 엑셀 업데이트 (전체 후보자 중 30점 이상만 제안문구 입력)
        update_excel_with_offers(EXCEL_FILE, candidates, MIN_SCORE)

    except ValueError as e:
        print(f"❌ 오류: {e}")
        print("\n💡 OpenAI API 키 설정 방법:")
        print("   1. 환경변수: export OPENAI_API_KEY_COMPANY='sk-...'")
        print("   2. 코드 수정: generator = PositionOfferGenerator(api_key='sk-...')")


if __name__ == "__main__":
    main()
