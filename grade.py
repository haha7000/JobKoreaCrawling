"""
인재 채용 평가 시스템 (Python 버전)
grade.js를 Python으로 변환
입력: kspac2022_with_introduction.json
출력: kspac2022_scored.json (점수 필터링된 결과)
"""
import json
import re
from pathlib import Path
from typing import Dict, List, Optional


class CandidateScorer:
    """채용 후보자 점수 계산기"""

    # 키워드 사전
    KW = {
        "insuranceSales": ["보험 영업", "보험영업", "보험상품", "설계사", "FP", "GA", "생명보험", "손해보험", "보장 분석", "보장설계", "종합재무설계", "리모델링"],
        "financeSales": ["금융 영업", "금융상품", "자산관리", "PB", "WM", "펀드", "증권", "투자", "대출상담", "카드영업", "지점 영업"],
        "generalSales": ["영업", "세일즈", "판매", "B2B 영업", "B2C 영업", "영업관리", "상담원", "상담", "텔레마케팅", "TM", "영업지원", "영업기획", "고객유치", "가망고객", "리드", "콜"],
        "indirectSales": ["고객응대", "CS", "시장조사", "프로모션", "홍보", "행사 운영", "매장관리", "판촉"],
        "activity": ["동아리", "인턴", "대외활동", "프로젝트", "공모전", "서포터즈", "홍보대사"],
        "commStrong": ["고객 니즈", "니즈 파악", "경청", "문제 해결", "클레임", "VOC", "고객 만족", "재구매", "추천", "관계 형성", "관계관리", "상담 스크립트", "컨설팅", "제안", "설득"],
        "commMedium": ["소통", "협업", "커뮤니케이션", "팀워크", "협력", "친화력", "긍정", "배려", "설명"],
        "certHigh": ["손해사정사", "AFPK", "CFP", "투자자산운용사", "증권투자권유대행인", "파생상품투자권유자문인력", "보험계리사"],
        "certBasic": ["보험 모집인", "생명보험 모집인", "손해보험 모집인", "펀드투자권유대행인", "펀드투자상담사", "은행FP", "퇴직연금", "신용분석사"],
        "certLight": ["운전면허", "2종보통", "1종보통", "CS리더스", "MOS"],
        "majorFinance": ["금융", "경제", "경영", "보험", "재무", "회계", "금융공학", "보험계리", "비즈니스"],
        "eduFinance": ["금융 교육", "펀드 교육", "자산관리 교육", "세일즈 교육", "세일즈 트레이닝", "상담 스킬", "세일즈 아카데미", "콜 교육", "FP 교육"],
        "motiveStrong": ["보험 산업", "보험업", "GA 채널", "모집질서", "준법", "소비자보호", "보장분석", "리드관리", "고객발굴", "리텐션", "리쿠르팅", "월납", "보장성", "인바운드/아웃바운드", "컨설팅영업"],
        "motiveWeak": ["성장", "열정", "도전", "문제 해결", "목표", "성과", "책임감", "자기계발"],
        "langNames": ["TOEIC", "토익", "OPIC", "OPIc", "오픽", "TOEFL", "IELTS"]
    }

    # 정규식 패턴
    RX = {
        "numberHit": re.compile(r'(\d{2,}\s*(%|건|명|회|개|만원|억|개월|주|일))|(\+\d{1,}%)', re.IGNORECASE),
        "toeic": re.compile(r'(TOEIC|토익)\s*[:\-]?\s*(\d{3,4})', re.IGNORECASE),
        "opic": re.compile(r'(OPIC|OPIc|오픽)\s*[:\-]?\s*([AIL]\w?)', re.IGNORECASE)
    }

    @staticmethod
    def normalize(text: str) -> str:
        """텍스트 정규화"""
        return str(text).lower().strip()

    @staticmethod
    def has_any(text: str, keywords: List[str]) -> bool:
        """텍스트에 키워드가 하나라도 있는지 확인"""
        normalized = CandidateScorer.normalize(text)
        for kw in keywords:
            if CandidateScorer.normalize(kw) in normalized:
                return True
        return False

    @staticmethod
    def extract_intro_text(intro_list: Optional[List[Dict]]) -> str:
        """자기소개서 텍스트 추출"""
        if not intro_list:
            return ""
        texts = []
        for item in intro_list:
            if isinstance(item, dict) and item.get("body_text"):
                texts.append(item["body_text"])
        return " ".join(texts)

    @staticmethod
    def extract_cert_text(cert_list: Optional[List[Dict]]) -> str:
        """자격증 텍스트 추출"""
        if not cert_list:
            return ""
        texts = []
        for item in cert_list:
            if isinstance(item, dict) and item.get("자격증명"):
                texts.append(item["자격증명"])
        return " ".join(texts)

    def score_sales_exp(self, candidate: Dict) -> int:
        """영업 경력 점수 (최대 30점)"""
        # 텍스트 수집
        exp_text = " ".join([
            str(candidate.get("경력", "")),
            str(candidate.get("제목", "")),
            str(candidate.get("직무", "")),
            str(candidate.get("기술스택", ""))
        ])

        # 활동 여부
        act_hit = self.has_any(exp_text, self.KW["activity"]) or self.RX["numberHit"].search(exp_text)

        # 기본 점수
        base = 0
        if self.has_any(exp_text, self.KW["insuranceSales"]) or self.has_any(exp_text, self.KW["financeSales"]):
            base = 20
        elif self.has_any(exp_text, self.KW["generalSales"]):
            base = 15
        elif self.has_any(exp_text, self.KW["indirectSales"]):
            base = 10

        # 활동 점수
        activity = 0
        if act_hit:
            activity = 5 if self.RX["numberHit"].search(exp_text) else 3

        exp_25 = min(25, base + activity)

        # 지원 분야 점수
        job_field = str(candidate.get("직무", ""))
        job_5 = 0
        if self.has_any(job_field, self.KW["insuranceSales"] + self.KW["financeSales"]):
            job_5 = 5
        elif self.has_any(job_field, self.KW["generalSales"]):
            job_5 = 3

        return exp_25 + job_5

    def score_customer_comm(self, candidate: Dict) -> int:
        """고객 커뮤니케이션 점수 (최대 5점) - 자기소개서 제거됨"""
        # 자기소개서 관련 점수는 0점
        comm_20 = 0

        # 인재상 점수 (없음)
        fit_5 = 0

        return comm_20 + fit_5

    def score_specialization(self, candidate: Dict) -> int:
        """전문성 점수 (최대 30점)"""
        # 자격증 텍스트
        cert_text = self.extract_cert_text(candidate.get("자격증"))
        cert_text += " " + str(candidate.get("기술스택", ""))

        # 학력 텍스트
        edu_text = str(candidate.get("학력", "")) + " " + str(candidate.get("제목", ""))

        # 자격증 점수
        cert_20 = 0
        if self.has_any(cert_text, self.KW["certHigh"]):
            cert_20 = 20
        elif self.has_any(cert_text, self.KW["certBasic"]):
            cert_20 = 10
        elif self.has_any(cert_text, self.KW["certLight"]):
            cert_20 = 2

        # 학력/교육 점수
        edu_10 = 0
        if self.has_any(edu_text, self.KW["majorFinance"]) or self.has_any(edu_text, self.KW["eduFinance"]):
            edu_10 = 10
        elif self.has_any(edu_text, ["경제원론", "재무회계", "마케팅", "금융상품"]):
            edu_10 = 5

        return cert_20 + edu_10

    def score_motivation_and_lang(self, candidate: Dict) -> int:
        """동기 및 어학 점수 (최대 5점) - 자기소개서 제거됨"""
        # 자기소개서 관련 점수는 0점
        mot_10 = 0

        # 어학 텍스트
        lang_text = str(candidate.get("기술스택", ""))

        # 어학 점수
        lang_5 = 0

        # TOEIC 확인
        toeic_match = self.RX["toeic"].search(lang_text)
        if toeic_match:
            toeic = int(toeic_match.group(2))
            if toeic >= 900:
                lang_5 = 5
            elif toeic >= 700:
                lang_5 = 3
            elif toeic >= 600:
                lang_5 = 1

        # OPIC 확인
        if lang_5 == 0:
            opic_match = self.RX["opic"].search(lang_text)
            if opic_match:
                opic_level = opic_match.group(2).upper()
                if re.search(r'AL|IH', opic_level):
                    lang_5 = 5
                elif 'IM' in opic_level:
                    lang_5 = 3
                else:
                    lang_5 = 1

        # 기타 어학 점수
        if lang_5 == 0 and self.has_any(lang_text, self.KW["langNames"]):
            lang_5 = 1

        return mot_10 + lang_5

    def calculate_score(self, candidate: Dict) -> Dict:
        """총점 계산"""
        sales_exp = self.score_sales_exp(candidate)
        customer_comm = self.score_customer_comm(candidate)
        specialization = self.score_specialization(candidate)
        motivation_lang = self.score_motivation_and_lang(candidate)

        total = sales_exp + customer_comm + specialization + motivation_lang

        return {
            "영업경력점수": sales_exp,
            "커뮤니케이션점수": customer_comm,
            "전문성점수": specialization,
            "동기어학점수": motivation_lang,
            "총점": total
        }


def update_excel_with_scores(excel_path: str, candidates: List[Dict]):
    """
    기존 엑셀 파일에 점수 컬럼 추가

    Args:
        excel_path: 엑셀 파일 경로
        candidates: 점수가 포함된 후보자 리스트
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment

    # 이력서번호 → 점수 매핑
    score_map = {}
    for candidate in candidates:
        rno = candidate.get("이력서번호")
        if rno and "점수상세" in candidate:
            score_map[str(rno)] = candidate["점수상세"]["총점"]

    # 엑셀 파일 열기
    if not Path(excel_path).exists():
        print(f"⚠️  엑셀 파일을 찾을 수 없습니다: {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 헤더 찾기
    headers = [cell.value for cell in ws[1]]

    # "점수" 컬럼이 이미 있는지 확인
    if "점수" in headers:
        print(f"⚠️  '점수' 컬럼이 이미 존재합니다. 기존 점수를 업데이트합니다.")
        score_col_idx = headers.index("점수") + 1
    else:
        # "나이" 컬럼 위치 찾기
        if "나이" not in headers:
            print(f"⚠️  '나이' 컬럼을 찾을 수 없습니다.")
            return

        age_col_idx = headers.index("나이") + 1  # openpyxl은 1-based index

        # "점수" 컬럼 삽입 (나이 다음)
        score_col_idx = age_col_idx + 1
        ws.insert_cols(score_col_idx)

        # 헤더 추가
        ws.cell(row=1, column=score_col_idx).value = "점수"
        ws.cell(row=1, column=score_col_idx).font = Font(bold=True)
        ws.cell(row=1, column=score_col_idx).alignment = Alignment(horizontal='center')

        # 헤더 다시 읽기 (컬럼 삽입 후)
        headers = [cell.value for cell in ws[1]]

    # 이력서번호 컬럼 위치 찾기
    rno_col_idx = None
    if "이력서번호" in headers:
        rno_col_idx = headers.index("이력서번호") + 1

    if not rno_col_idx:
        print(f"⚠️  '이력서번호' 컬럼을 찾을 수 없습니다.")
        return

    # 각 행에 점수 입력
    updated_count = 0
    for row_idx in range(2, ws.max_row + 1):
        rno = ws.cell(row=row_idx, column=rno_col_idx).value
        if rno and str(rno) in score_map:
            score = score_map[str(rno)]
            ws.cell(row=row_idx, column=score_col_idx).value = score
            ws.cell(row=row_idx, column=score_col_idx).alignment = Alignment(horizontal='center')
            updated_count += 1

    # 점수 컬럼 너비 조정
    ws.column_dimensions[openpyxl.utils.get_column_letter(score_col_idx)].width = 10

    # 저장
    wb.save(excel_path)
    print(f"\n📊 엑셀 파일 업데이트 완료!")
    print(f"   파일: {excel_path}")
    print(f"   점수 입력: {updated_count}개 행")


def grade_candidates(
    input_json: str,
    output_json: str,
    excel_path: str = None,
    min_score: int = 30
):
    """
    후보자 채점 및 필터링

    Args:
        input_json: 입력 JSON 파일 (kspac2022_with_introduction.json)
        output_json: 출력 JSON 파일 (점수 필터링된 결과)
        excel_path: 엑셀 파일 경로 (점수 컬럼 추가용)
        min_score: 최소 합격 점수 (기본 30점)
    """
    # 입력 파일 로드
    if not Path(input_json).exists():
        print(f"❌ 파일을 찾을 수 없습니다: {input_json}")
        return

    with open(input_json, 'r', encoding='utf-8') as f:
        candidates = json.load(f)

    print(f"📋 총 {len(candidates)}명 채점 시작")
    print(f"🎯 합격 기준: {min_score}점 이상\n")

    scorer = CandidateScorer()
    passed_candidates = []
    failed_count = 0
    all_candidates_with_score = []  # 모든 후보자 (엑셀 업데이트용)

    # 각 후보자 채점
    for idx, candidate in enumerate(candidates, 1):
        name = candidate.get("이름", "Unknown")

        # 점수 계산
        scores = scorer.calculate_score(candidate)
        total_score = scores["총점"]

        # 점수 추가
        candidate["점수상세"] = scores
        all_candidates_with_score.append(candidate)

        # 합격/불합격 판정
        if total_score >= min_score:
            passed_candidates.append(candidate)
            print(f"[{idx}/{len(candidates)}] ✅ {name} - {total_score}점 (합격)")
        else:
            failed_count += 1
            print(f"[{idx}/{len(candidates)}] ❌ {name} - {total_score}점 (불합격)")

    # 점수 순으로 정렬
    passed_candidates.sort(key=lambda x: x["점수상세"]["총점"], reverse=True)

    # 결과 저장 (합격자만)
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(passed_candidates, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"✅ 채점 완료!")
    print(f"   총 인원: {len(candidates)}명")
    print(f"   합격: {len(passed_candidates)}명")
    print(f"   불합격: {failed_count}명")
    print(f"   합격률: {len(passed_candidates)/len(candidates)*100:.1f}%")
    print(f"\n💾 저장: {output_json}")
    print(f"{'='*60}")

    # 엑셀 파일 업데이트 (모든 후보자 점수 추가)
    if excel_path:
        update_excel_with_scores(excel_path, all_candidates_with_score)

    # 원본 JSON 파일도 점수 업데이트 (position_offer.py에서 사용)
    with open(input_json, 'w', encoding='utf-8') as f:
        json.dump(all_candidates_with_score, f, ensure_ascii=False, indent=2)
    print(f"\n💾 원본 JSON 업데이트: {input_json} (점수 포함)")


def main():
    """메인 실행"""
    INPUT_FILE = "output/kspac2022_with_introduction.json"
    OUTPUT_FILE = "output/kspac2022_scored.json"
    EXCEL_FILE = "output/kspac2022_결과.xlsx"  # main.py에서 생성된 엑셀 파일
    MIN_SCORE = 30  # 최소 합격 점수

    grade_candidates(INPUT_FILE, OUTPUT_FILE, EXCEL_FILE, MIN_SCORE)


if __name__ == "__main__":
    main()
